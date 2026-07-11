"""Parse management consolidated Excel into storable matrix rows."""
from __future__ import annotations

import re
from pathlib import Path
from typing import Any

COLUMN_CODES = ["AH", "AK", "AM", "MM", "BIU", "NP", "PV", "RT", "SP", "VC", "VP", "VS"]

PLAN_LABEL_MAP = {
    "green": "green",
    "amber": "amber",
    "bluesky": "bluesky",
    "blue sky": "bluesky",
    "total": "total",
}
BIFUR_LABEL_MAP = {
    "unidentified bluesky": "unidentified",
    "known bluesky": "known",
    "total bluesky": "total_bs",
}

MONTH_NAME_TO_KEY = {
    "january": "01", "february": "02", "march": "03", "april": "04",
    "may": "05", "june": "06", "july": "07", "august": "08",
    "september": "09", "october": "10", "november": "11", "december": "12",
    "jan": "01", "feb": "02", "mar": "03", "apr": "04",
    "jun": "06", "jul": "07", "aug": "08", "sep": "09", "oct": "10", "nov": "11", "dec": "12",
}


def _slugify(text: str) -> str:
    s = re.sub(r"[^a-z0-9]+", "_", text.lower()).strip("_")
    return s[:120] or "row"


def _fy_slug_from_text(text: str) -> str | None:
    m = re.search(r"fy\s*(\d{2})\s*[-–]\s*(\d{2})", text, re.I)
    if m:
        return f"{m.group(1)}{m.group(2)}"
    m = re.search(r"(\d{2})\s*[-–]\s*(\d{2})", text)
    if m and int(m.group(1)) < 50:
        return f"{m.group(1)}{m.group(2)}"
    return None


def _month_key_from_text(text: str) -> str | None:
    lower = text.lower()
    if "early april" in lower or re.search(r"\bapril\b", lower):
        return "04"
    for name, key in MONTH_NAME_TO_KEY.items():
        if re.search(rf"\b{name}\b", lower):
            return key
    return None


def _tone_from_label(label: str) -> str | None:
    ll = label.lower().strip()
    if ll == "green":
        return "green"
    if ll == "amber":
        return "amber"
    if ll in ("bluesky", "blue sky") or "bluesky" in ll:
        return "bluesky"
    return None


def _parse_section_context(label: str) -> dict[str, Any]:
    lower = label.lower()
    fy = _fy_slug_from_text(label) or "2627"

    if "bifurcation of bluesky" in lower:
        mk = _month_key_from_text(label)
        return {"fy": fy, "context": f"bifur_{mk}" if mk else "bifur", "section_label": label}

    if "details as per initial plan" in lower:
        return {"fy": fy, "context": "initial", "section_label": label}
    if "business plan decided by board" in lower or "business plan decided by the board" in lower:
        return {"fy": fy, "context": "board", "section_label": label}
    if "business plan update" in lower:
        mk = _month_key_from_text(label)
        return {"fy": fy, "context": f"monthly_{mk}" if mk else "monthly", "section_label": label}
    if "summary of collections" in lower:
        return {"fy": fy, "context": "coll_summary", "section_label": label}

    return {"fy": fy, "context": None, "section_label": label}


def _standalone_row_key(label: str, report_fy: str) -> str:
    lower = label.lower().strip()
    if lower.startswith("projected plan"):
        return "hist_fy2425_projected"
    if "actual collections (fy 24-25)" in lower:
        return "hist_fy2425_actual"
    if "actual collections (fy 25-26)" in lower:
        return "hist_fy2526_actual"
    if "variance in collections" in lower and "25-26" in lower:
        return "hist_fy2526_coll_variance"
    if "bluesky achieved" in lower:
        mk = _month_key_from_text(label)
        return f"hist_bluesky_achieved_{mk}" if mk else _slugify(label)
    if "planned collection" in lower:
        mk = _month_key_from_text(label)
        suffix = "_revised" if "revised" in lower else ""
        return f"fy{report_fy}_coll_planned_{mk}{suffix}" if mk else _slugify(label)
    if "actual collection" in lower and "collections -" not in lower:
        mk = _month_key_from_text(label)
        return f"fy{report_fy}_coll_actual_{mk}" if mk else _slugify(label)
    if lower.startswith("collections upto"):
        mk = _month_key_from_text(label)
        return f"fy{report_fy}_coll_upto_{mk}" if mk else _slugify(label)
    if lower.startswith("actual collections -"):
        mk = _month_key_from_text(label)
        return f"fy{report_fy}_coll_summary_actual_{mk}" if mk else _slugify(label)
    if lower.startswith("expected collections"):
        mk = _month_key_from_text(label)
        return f"fy{report_fy}_coll_summary_expected_{mk}" if mk else _slugify(label)
    if "total collections (actual" in lower:
        return f"fy{report_fy}_coll_summary_total"
    return _slugify(label)


def _data_row_key(label: str, ctx: dict[str, Any] | None, report_fy: str) -> str:
    lower = label.lower().strip()
    if not ctx or not ctx.get("context"):
        return _standalone_row_key(label, report_fy)

    fy = ctx.get("fy") or report_fy
    context = ctx["context"]

    if context.startswith("bifur"):
        part = BIFUR_LABEL_MAP.get(lower)
        if part:
            return f"fy{fy}_{context}_{part}"
        return f"fy{fy}_{context}_{_slugify(label)}"

    part = PLAN_LABEL_MAP.get(lower)
    if part:
        return f"fy{fy}_{context}_{part}"
    return f"fy{fy}_{context}_{_slugify(label)}"


def _values_empty(values: dict[str, Any]) -> bool:
    return all(v is None or v == "" for v in values.values())


def _read_row_values(ws, row_idx: int) -> tuple[dict[str, float | None], float | None]:
    values: dict[str, float | None] = {}
    for i, code in enumerate(COLUMN_CODES):
        raw = ws.cell(row_idx, i + 2).value
        if raw is None or raw == "":
            values[code] = None
        else:
            try:
                values[code] = float(raw)
            except (TypeError, ValueError):
                values[code] = None
    total_raw = ws.cell(row_idx, 14).value
    total = float(total_raw) if total_raw not in (None, "") else None
    return values, total


def parse_consolidated_xlsx(path: str | Path, report_fy: str = "2627") -> list[dict[str, Any]]:
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Summary"] if "Summary" in wb.sheetnames else wb.active

    rows_out: list[dict[str, Any]] = []
    ctx: dict[str, Any] | None = None
    sort_order = 0

    for r in range(5, ws.max_row + 1):
        label = ws.cell(r, 1).value
        if not label or not str(label).strip():
            continue
        label = str(label).strip()
        values, imported_total = _read_row_values(ws, r)

        ll = label.lower()
        is_plan_label = ll in PLAN_LABEL_MAP or ll in BIFUR_LABEL_MAP

        if not is_plan_label and _values_empty(values):
            if ll.startswith("bifurcation of bluesky"):
                rows_out.append({"kind": "subheader", "label": label, "sort_order": sort_order})
                sort_order += 1
                ctx = _parse_section_context(label)
                continue
            rows_out.append({"kind": "section", "label": label, "sort_order": sort_order})
            sort_order += 1
            ctx = _parse_section_context(label)
            continue

        row_key = _data_row_key(label, ctx, report_fy)
        rows_out.append({
            "kind": "data",
            "label": label,
            "row_key": row_key,
            "tone": _tone_from_label(label),
            "values": values,
            "imported_total": imported_total,
            "sort_order": sort_order,
        })
        sort_order += 1

    return rows_out


def default_xlsx_path() -> Path:
    root = Path(__file__).resolve().parents[3]
    candidates = [
        root / "csv-templates" / "filled" / "Business Plan_Consolidated_FY 2026-27.xlsx",
        root / "backend" / "csv" / "Business Plan_Consolidated_FY 2026-27.xlsx",
    ]
    for p in candidates:
        if p.exists():
            return p
    return candidates[0]
